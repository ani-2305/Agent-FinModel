# render_excel_model.py

from openpyxl import Workbook # type: ignore 
from openpyxl.utils.dataframe import dataframe_to_rows # type: ignore
from . import shared_state
from .format_excel_sheets import format_excel_sheets
from .apply_conditional_formatting import apply_conditional_formatting
from .error_handling import handle_error

def render_excel_model(template_path: str, output_path: str) -> dict:
    """
    Renders the financial model into an Excel file using a template.

    :param template_path: Path to the Excel template file.
    :param output_path: Path to save the rendered Excel model.
    :return: Success message or an error message.
    """
    try:
        # Load the workbook
        from openpyxl import load_workbook # type: ignore
        wb = load_workbook(template_path)

        # Get the sheets
        income_sheet = wb['Forecast Income Statement']
        balance_sheet = wb['Forecast Balance Sheet']
        cash_flow_sheet = wb['Forecast Cash Flow Statement']

        # Get forecasted data
        forecast_income = shared_state.model_data['forecast_income_statement']
        forecast_balance = shared_state.model_data['forecast_balance_sheet']
        forecast_cash_flow = shared_state.model_data.get('forecast_cash_flow_statement')

        # Write data to sheets
        for r_idx, row in enumerate(dataframe_to_rows(forecast_income, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                income_sheet.cell(row=r_idx, column=c_idx, value=value)

        for r_idx, row in enumerate(dataframe_to_rows(forecast_balance, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                balance_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Similar steps for Cash Flow Statement if available
        # ...

        # Apply formatting
        format_excel_sheets(wb)
        apply_conditional_formatting(wb)

        # Save the workbook
        wb.save(output_path)

        return {"message": f"Excel model successfully saved to {output_path}"}

    except Exception as e:
        return {"error": handle_error(e)}

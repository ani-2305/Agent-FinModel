# apply_conditional_formatting.py

from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from .error_handling import handle_error

def apply_conditional_formatting(workbook) -> dict:
    """
    Applies conditional formatting to highlight key data points.

    :param workbook: An openpyxl Workbook object.
    :return: Success message or an error message.
    """
    try:
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for sheet_name in ['Forecast Income Statement', 'Forecast Balance Sheet']:
            ws = workbook[sheet_name]
            # Apply conditional formatting to highlight negative numbers
            dxf = DifferentialStyle(fill=red_fill)
            rule = Rule(type='cellIs', operator='lessThan', formula=['0'], dxf=dxf)
            ws.conditional_formatting.add('A1:Z1000', rule)

        return {"message": "Conditional formatting successfully applied."}

    except Exception as e:
        return {"error": handle_error(e)}

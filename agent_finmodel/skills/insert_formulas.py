# insert_formulas.py

from openpyxl import Workbook # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from . import shared_state
from .error_handling import handle_error

def insert_formulas(workbook: Workbook) -> dict:
    """
    Inserts Excel formulas into the workbook for dynamic calculations.

    :param workbook: An openpyxl Workbook object.
    :return: Success message or an error message.
    """
    try:
        ws = workbook['Forecast Income Statement']
        max_row = ws.max_row
        max_col = ws.max_column

        # Insert formula for Gross Profit
        for row in range(2, max_row + 1):
            revenue_cell = ws.cell(row=row, column=2).coordinate
            cogs_cell = ws.cell(row=row, column=3).coordinate
            gross_profit_cell = ws.cell(row=row, column=4)
            gross_profit_cell.value = f"={revenue_cell}-{cogs_cell}"

        # Similar formula insertions for other calculated fields
        # ...

        return {"message": "Formulas successfully inserted into the workbook."}

    except Exception as e:
        return {"error": handle_error(e)}
